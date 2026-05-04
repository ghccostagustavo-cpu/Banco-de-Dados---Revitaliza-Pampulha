# =============================================================================
# IMPORTAÇÃO DE BIBLIOTECAS
# =============================================================================

import pandas as pd
import sqlite3
from datetime import datetime

# =============================================================================
# CONFIGURAÇÕES DE CAMINHOS E TABELAS
# =============================================================================

# Caminho para a planilha de cadastro técnico/geográfico
caminho_cadastro = r"caminho\para\cadastro.xlsx"
tabela_cadastro = "cadastro"
aba_cadastro = "aba_cadastro"

# Definição dos destinos de saída: banco de dados e registros de atividade
caminho_banco = r"caminho_banco.db"
caminho_log = r"caminho_log.txt"


def registrar_log(msg):
    """
    Gera um histórico de execução com data e hora. 
    Salva no arquivo de log e exibe no console.
    """
    agora = datetime.now().strftime("%d-%m-%Y, %H:%M:%S")
    linha = f"[{agora}] {msg}\n"
    # Abre o log com encoding específico para evitar erros de caracteres especiais
    with open(caminho_log, "a", encoding="ISO-8859-1") as arquivo_log:
        arquivo_log.write(linha)
    print(linha.strip())


def limpar_converter(df):
    """
    Realiza o tratamento pesado: remove colunas desnecessárias por intervalo de letras
    e converte coordenadas geográficas para formato decimal.
    """
    from openpyxl.utils import get_column_letter, column_index_from_string
    import openpyxl
    import re

    # Função interna para transformar intervalos (ex: 'A' até 'C') em uma lista de letras
    def letras_range(inicio, fim):
        return [get_column_letter(i) for i in range(
            column_index_from_string(inicio),
            column_index_from_string(fim) + 1
        )]

    # Mapeamento de todas as colunas que devem ser removidas da planilha original
    excluir_colunas = set(
        letras_range('A', 'C') +
        letras_range('H', 'H') +
        letras_range('N', 'N') +
        letras_range('P', 'X') +
        letras_range('Z', 'AD') +
        letras_range('AF', 'FK') +
        letras_range('FN', 'FO')
    )

    # Abre o arquivo via openpyxl apenas para ler os nomes das colunas no cabeçalho (linha 6)
    wb = openpyxl.load_workbook(caminho_cadastro, read_only=True)
    ws = wb[aba_cadastro]
    cabecalho = {cell.column_letter: cell.value for cell in list(ws.rows)[5]}
    wb.close()

    # Cria uma lista de nomes de colunas baseada nas letras mapeadas anteriormente
    nomes_excluir = [
        cabecalho[l] for l in excluir_colunas
        if l in cabecalho and cabecalho[l]
    ]

    registrar_log(f"Excluindo {len(nomes_excluir)} colunas desnecessárias...")

    # Remove as colunas identificadas do DataFrame principal
    df = df.drop(columns=[col for col in nomes_excluir if col in df.columns])

    def dms_para_decimal(valor):
        """
        Converte coordenadas no formato 23° 15' 45" S para -23.2625000
        """
        try:
            # Extrai apenas números do texto usando Expressão Regular (Regex)
            nums = re.findall(r'[\d,.]+', str(valor))
            graus = float(nums[0])
            minutos = float(nums[1])
            # Garante que o separador decimal seja ponto para o cálculo
            segundos = float(nums[2].replace(',', '.'))
            
            # Fórmula matemática de conversão: Graus + (Minutos/60) + (Segundos/3600)
            decimal = graus + minutos / 60 + segundos / 3600
            
            # Se for Sul ou Oeste, a coordenada decimal deve ser negativa
            if 'S' in str(valor) or 'O' in str(valor) or 'W' in str(valor):
                decimal = -decimal
            return round(decimal, 7)
        except:
            return None

    # Aplica a conversão nas colunas de Latitude e Longitude
    df['Latitude_decimal'] = df['Latitude'].apply(dms_para_decimal)
    df['Longitude_decimal'] = df['Longitude'].apply(dms_para_decimal)
    
    # Cria uma coluna no formato WKT (Well-Known Text) usada por softwares de GIS (ex: QGIS)
    df['Coordenadas'] = 'POINT(' + df['Longitude_decimal'].astype(str) + ' ' + df['Latitude_decimal'].astype(str) + ')'

    return df


def sincronizar_dados_cadastro():
    """
    Fluxo principal: lê o Excel a partir da linha 6 (header=5), trata e salva no SQLite.
    """
    try:
        registrar_log("Lendo cadastro...")
        # header=5 indica que o título das colunas está na 6ª linha da planilha
        df = pd.read_excel(caminho_cadastro, sheet_name=aba_cadastro, header=5, engine='openpyxl')

        registrar_log("Limpando e convertendo dados...")
        df = limpar_converter(df)

        registrar_log("Conectando ao banco SQLite...")
        conexao = sqlite3.connect(caminho_banco)
        # Salva os dados processados na tabela definida, substituindo se já existir
        df.to_sql(tabela_cadastro, conexao, if_exists='replace', index=False)
        conexao.close()

        registrar_log(f"SUCESSO: {len(df)} linhas sincronizadas para a tabela '{tabela_cadastro}'.")

    except Exception as e:
        registrar_log(f"ERRO: {str(e)}")


if __name__ == "__main__":
    sincronizar_dados_cadastro()
    print("Sincronização concluída com sucesso!")